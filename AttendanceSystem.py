# Tensorflow version == 2.0.0
import numpy as np
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import Sequential,Model
from tensorflow.keras.layers import ZeroPadding2D,Convolution2D,MaxPooling2D
from tensorflow.keras.layers import Dense,Dropout,Softmax,Flatten,Activation,BatchNormalization
from tensorflow.keras.preprocessing.image import load_img,img_to_array
from tensorflow.keras.applications.imagenet_utils import preprocess_input
import tensorflow.keras.backend as K
import os, sys
import cv2
import _sqlite3
from matplotlib import pyplot as plt
import openpyxl

wbkName = 'AttendanceSheet.xlsx'
wbk = openpyxl.load_workbook(wbkName)

def Mark_Attendance(name, currentcol):
   for wks in wbk.worksheets:
      for myRow in range(1, 10):
         if wks.cell(myRow,1).value==name:
            wks.cell(myRow, currentcol).value = 'P'

   wbk.save(wbkName)
# VGG_FACE architecture
model = Sequential()    #defines ann model
#Defining 16 convolutional layers for VGG
model.add(ZeroPadding2D((1,1),input_shape=(224,224, 3)))  #input layer for images
model.add(Convolution2D(64, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(64, (3, 3), activation='relu'))
model.add(MaxPooling2D((2,2), strides=(2,2)))
model.add(ZeroPadding2D((1,1)))	
model.add(Convolution2D(128, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(128, (3, 3), activation='relu'))
model.add(MaxPooling2D((2,2), strides=(2,2)))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(256, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(256, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(256, (3, 3), activation='relu'))
model.add(MaxPooling2D((2,2), strides=(2,2)))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(MaxPooling2D((2,2), strides=(2,2)))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(ZeroPadding2D((1,1)))
model.add(Convolution2D(512, (3, 3), activation='relu'))
model.add(MaxPooling2D((2,2), strides=(2,2)))
model.add(Convolution2D(4096, (7, 7), activation='relu'))
model.add(Dropout(0.5))
model.add(Convolution2D(4096, (1, 1), activation='relu'))
model.add(Dropout(0.5))
model.add(Convolution2D(2622, (1, 1)))
model.add(Flatten())
model.add(Activation('softmax'))

# Load VGG Face model weights
model.load_weights('vgg_face_weights.h5') #where VGG weights of trained data are placed
#model.summary()  #gives summary of model
 
vgg_face=Model(inputs=model.layers[0].input,outputs=model.layers[-2].output) 





#Data to train the model
x_train=[]
y_train=[]
persons=dict() #creating persons dictionary
#The following directory has the training dataset, the name of each folder is the name of person

person_folders=os.listdir('C:\\Users\\Admin\\Desktop\\Face recognition\\AttendanceSystem\\Training_images') 
for i,person in enumerate(person_folders):
  persons[i]=person
  images=os.listdir('Training_images/'+person+'/')
  for image in images:
    img=load_img('Training_images/'+person+'/'+image,target_size=(224,224))
    img=img_to_array(img)
    img=np.expand_dims(img,axis=0)
    img=preprocess_input(img)
    img_encode=vgg_face(img)
    x_train.append(np.squeeze(K.eval(img_encode)).tolist())
    y_train.append(i)
#print(persons) 
# Prepare Test Data
x_test=[]
y_test=[]
persons=dict()
person_folders=os.listdir('C:\\Users\\Admin\\Desktop\\Face recognition\\AttendanceSystem\\Testing_Images')
for i,person in enumerate(person_folders):
  images=os.listdir('Testing_Images/'+person+'/')
  for image in images:
    img=load_img('Testing_Images/'+person+'/'+image,target_size=(224,224))  #224X224 dimensions of image for VGG
    img=img_to_array(img)
    img=np.expand_dims(img,axis=0)
    img=preprocess_input(img)
    img_encode=vgg_face(img)
    x_test.append(np.squeeze(K.eval(img_encode)).tolist())
    y_test.append(i)




x_train=np.array(x_train) 
y_train=np.array(y_train)
x_test=np.array(x_test) 
y_test=np.array(y_test) 



#Saving test and training data
np.save('train_data',x_train)
np.save('train_labels',y_train)
np.save('test_data',x_test)
np.save('test_labels',y_test)


#loading saved data
x_train=np.load('train_data.npy')
y_train=np.load('train_labels.npy')
x_test=np.load('test_data.npy')
y_test=np.load('test_labels.npy')


#Training softmax classifier
classifier_model=Sequential()	
classifier_model.add(Dense(units=100,input_dim=x_train.shape[1],kernel_initializer='glorot_uniform'))		
classifier_model.add(BatchNormalization())		
classifier_model.add(Activation('tanh'))
classifier_model.add(Dropout(0.3))
classifier_model.add(Dense(units=10,kernel_initializer='glorot_uniform'))
classifier_model.add(BatchNormalization())
classifier_model.add(Activation('tanh'))
classifier_model.add(Dropout(0.2))
classifier_model.add(Dense(units=7,kernel_initializer='he_uniform')) #units has the number of inpit images no+1
classifier_model.add(Activation('softmax'))
classifier_model.compile(loss=tf.keras.losses.SparseCategoricalCrossentropy(),optimizer='nadam',metrics=['accuracy'])


classifier_model.fit(x_train,y_train,epochs=100,validation_data=(x_test,y_test))


# Saving model
tf.keras.models.save_model(classifier_model,'face_classifier_model.h5')

# Load saved model
classifier_model=tf.keras.models.load_model('face_classifier_model.h5')

#Path for testing images
test_images_path='Test_Images'


def plot(img):
  plt.figure(figsize=(8,4))
  plt.imshow(img[:,:,::-1])
  plt.show()


  #defining labels for classes
persons={ 0: 'Dure', 1:'halsey',2:'harry',3:'liam',4:'louis',5:'niall',6:'zayn'}

#the folder where the recognized images are placed
os.mkdir('Recognized')


for wks in wbk.worksheets:
   current_col=wks.cell(80,1).value
   break


for img_name in os.listdir('Test_Images'):
  if img_name=='crop_img.jpg':
    continue
  # Load Image
  img=cv2.imread('Test_Images/'+img_name)


  black_image=cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
  faceCascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
  faces = faceCascade.detectMultiScale(black_image,1.3,3,minSize=(30, 30))
  print("Found"+str(len(faces))+"Faces!")
  for (x, y, w, h) in faces:
      cv2.rectangle(img, (x, y), (x+w, y+h), (0, 0, 255), 2)
      Region_of_Interest = img[y:y + h, x:x + w]
    
      cv2.imwrite('Test_Images/crop_img.jpg', Region_of_Interest)

      crop_img=load_img('Test_Images/crop_img.jpg',target_size=(224,224))
      crop_img=img_to_array(crop_img)
      crop_img=np.expand_dims(crop_img,axis=0)
      crop_img=preprocess_input(crop_img)
      img_encode=vgg_face(crop_img)

    #Making Predictions
      embed=K.eval(img_encode)
      person=classifier_model.predict(embed)
      name=persons[np.argmax(person)]
      os.remove('Test_Images/crop_img.jpg')
      cv2.rectangle(img, (x, y), (x+w, y+h), (0, 0, 255), 2)
      img=cv2.putText(img,name,(x,x+w-10),cv2.FONT_HERSHEY_SIMPLEX,1,(255,0,255),2,cv2.LINE_AA)
      Mark_Attendance(name, current_col)
      img=cv2.putText(img,str(np.max(person)),(x,x+10),cv2.FONT_HERSHEY_SIMPLEX,0.5,(0,0,255),1,cv2.LINE_AA)
  # Save images with bounding box,name and accuracy 
  cv2.imwrite('Recognized/'+img_name,img)
  plot(img)

for wks in wbk.worksheets:
  wks.cell(80,1).value = wks.cell(80,1).value + 1
  break

wbk.save(wbkName)
wbk.close

print('Attendance saved in Attendancesheet.xlsx!')






















